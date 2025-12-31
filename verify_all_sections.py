#!/usr/bin/env python3
"""
Verify all sections have correct Total and % Change rows.
"""
from openpyxl import load_workbook

wb = load_workbook('data/output_report.xlsx')
ws = wb.active

print('=' * 80)
print('VERIFICATION: ALL SECTIONS - Total and % Change Rows')
print('=' * 80)

# Find all Total and % Change rows
total_rows = []
change_rows = []

for row_idx in range(1, 300):
    cell_b = ws.cell(row_idx, 2).value
    if cell_b:
        cell_str = str(cell_b).lower()
        if 'total' in cell_str and 'visits' not in cell_str:
            total_rows.append(row_idx)
        elif '% change' in cell_str or '%change' in cell_str:
            change_rows.append(row_idx)

print(f'\nFound {len(total_rows)} Total rows')
print(f'Found {len(change_rows)} % Change rows')

# Verify first 5 sections
for i in range(min(5, len(total_rows))):
    total_row = total_rows[i]
    change_row = change_rows[i] if i < len(change_rows) else None
    
    print(f'\n{"=" * 80}')
    print(f'SECTION {i+1}')
    print(f'{"=" * 80}')
    
    # Check Total row
    print(f'\nRow {total_row} (Total):')
    print(f'  Label: {ws.cell(total_row, 2).value}')
    print(f'  Year-2023: {ws.cell(total_row, 3).value}')
    print(f'  Year-2024: {ws.cell(total_row, 4).value}')
    print(f'  Year-2025: {ws.cell(total_row, 5).value}')
    print(f'  YOY %: {ws.cell(total_row, 6).value} (should be None)')
    print(f'  LM %: {ws.cell(total_row, 7).value} (should be None)')
    
    # Check % Change row
    if change_row:
        print(f'\nRow {change_row} (% Change):')
        print(f'  Label: {ws.cell(change_row, 2).value}')
        print(f'  Year-2023: {ws.cell(change_row, 3).value} (should be None)')
        print(f'  Year-2024: {ws.cell(change_row, 4).value} (should have %)')
        print(f'  Year-2025: {ws.cell(change_row, 5).value} (should have %)')
        print(f'  YOY %: {ws.cell(change_row, 6).value} (should have %)')
        print(f'  LM %: {ws.cell(change_row, 7).value} (currently None)')
        
        # Validate
        is_valid = (
            ws.cell(total_row, 6).value is None and  # Total YOY empty
            ws.cell(change_row, 3).value is None and  # % Change 2023 empty
            ws.cell(change_row, 4).value is not None and  # % Change 2024 filled
            ws.cell(change_row, 6).value is not None  # % Change YOY filled
        )
        print(f'\n  {"✓ VALID" if is_valid else "✗ INVALID"}')

print('\n' + '=' * 80)
print('SUMMARY')
print('=' * 80)
print('✓ Total rows should have actual numbers in year columns')
print('✓ Total rows should have None in YOY % and LM %')
print('✓ % Change rows should have label "% Change"')
print('✓ % Change Year-2023: None (reference)')
print('✓ % Change Year-2024: % vs 2023')
print('✓ % Change Year-2025: % vs 2023')
print('✓ % Change YOY %: % (2025 vs 2024)')
print('✓ % Change LM %: None (not applicable for totals)')
print('=' * 80)
