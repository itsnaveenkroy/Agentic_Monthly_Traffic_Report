from openpyxl import load_workbook

# Check OUTPUT file
output_path = 'data/output_report.xlsx'
wb = load_workbook(output_path)
ws = wb.active

print('=' * 70)
print('TOTAL ROW VERIFICATION (Should have YOY/LM values, not 0)')
print('=' * 70)

# Find first Total row
for row in range(20, 26):
    cell_b = ws.cell(row, 2).value
    if cell_b and 'total' in str(cell_b).lower():
        print(f'\nRow {row}: {cell_b}')
        print(f'  Year 2023: {ws.cell(row, 3).value}')
        print(f'  Year 2024: {ws.cell(row, 4).value}')  
        print(f'  Year 2025: {ws.cell(row, 5).value}')
        print(f'  YOY %: {ws.cell(row, 6).value}')
        print(f'  LM %: {ws.cell(row, 7).value}')
        break

print('\n' + '=' * 70)
print('SUMMARY FORMATTING VERIFICATION (Rich text with colors)')
print('=' * 70)

summary_cell = ws.cell(11, 9)
print(f'\nSummary cell (I11) value type: {type(summary_cell.value).__name__}')

if 'CellRichText' in str(type(summary_cell.value)):
    print('✓ Rich text formatting DETECTED!')
    print(f'\nPreview: {str(summary_cell.value)[:300]}...')
else:
    print('✗ Plain text (no rich formatting)')
    if summary_cell.value:
        print(f'\nPreview: {str(summary_cell.value)[:300]}...')

print('\n' + '=' * 70)
print('DONE')
print('=' * 70)
