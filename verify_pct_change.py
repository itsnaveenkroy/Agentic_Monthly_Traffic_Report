#!/usr/bin/env python3
"""
Verify % Change row behavior is correct.
This % Change behavior is intentional. First column is a reference baseline.
"""
from openpyxl import load_workbook
import os

if not os.path.exists('data/output_report.xlsx'):
    print('❌ Output file does not exist. Run: python main.py')
    exit(1)

wb = load_workbook('data/output_report.xlsx')
ws = wb.active

print('=' * 70)
print('% CHANGE ROW VERIFICATION')
print('=' * 70)

# Check first % Change row (row 24)
row = 24
print(f'\nRow {row} (% Change):')
print(f'  Column B (Month): {ws.cell(row, 2).value}')
print(f'  Column C (Year-2023): {ws.cell(row, 3).value}')
print(f'  Column D (Year-2024): {ws.cell(row, 4).value}')
print(f'  Column E (Year-2025): {ws.cell(row, 5).value}')
print(f'  Column F (YOY %): {ws.cell(row, 6).value}')
print(f'  Column G (LM %): {ws.cell(row, 7).value}')

# Validate
year_2023_empty = ws.cell(row, 3).value is None
year_2024_empty = ws.cell(row, 4).value is None
year_2025_empty = ws.cell(row, 5).value is None
yoy_filled = ws.cell(row, 6).value is not None
lm_empty = ws.cell(row, 7).value is None

print('\n' + '=' * 70)
print('VALIDATION RESULTS')
print('=' * 70)
print(f'✓ Year-2023 EMPTY (reference): {year_2023_empty}')
print(f'✓ Year-2024 EMPTY: {year_2024_empty}')
print(f'✓ Year-2025 EMPTY: {year_2025_empty}')
print(f'✓ YOY % FILLED: {yoy_filled}')
print(f'✓ LM % EMPTY: {lm_empty}')

all_valid = year_2023_empty and year_2024_empty and year_2025_empty and yoy_filled and lm_empty

print('\n' + '=' * 70)
if all_valid:
    print('✅ PASS - % Change row behavior is CORRECT')
else:
    print('❌ FAIL - % Change row behavior is INCORRECT')
print('=' * 70)
