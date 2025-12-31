#!/usr/bin/env python3
"""
Check if % Change row should have LM % calculation
"""
from openpyxl import load_workbook

# Check output
print('=' * 80)
print('CHECKING LM % IN % CHANGE ROW')
print('=' * 80)

wb_out = load_workbook('data/output_report.xlsx')
ws_out = wb_out.active

print('\nOutput file % Change row (Row 24):')
for col in range(2, 10):
    header = ws_out.cell(11, col).value
    value = ws_out.cell(24, col).value
    print(f'  Col {col} ({header}): {value}')

# Check if input file exists to compare
try:
    wb_in = load_workbook('data/input_ga4_data.xlsx', data_only=True)
    ws_in = wb_in.active
    
    print('\n' + '=' * 80)
    print('CHECKING INPUT FILE FOR REFERENCE')
    print('=' * 80)
    
    # Find % Change rows in input
    for row in range(1, 300):
        cell = ws_in.cell(row, 2).value
        if cell and ('% change' in str(cell).lower() or '%change' in str(cell).lower()):
            print(f'\nFound % Change row at {row}:')
            for col in range(2, 10):
                value = ws_in.cell(row, col).value
                print(f'  Col {col}: {value}')
            break
    
except Exception as e:
    print(f'\nCould not check input file: {e}')

print('\n' + '=' * 80)
print('ANALYSIS')
print('=' * 80)
print('LM % (Last Month) is a month-over-month comparison.')
print('For the % Change row (which shows year totals), LM % does not apply.')
print('% Change row should show:')
print('  - Year columns: % vs reference year (2023)')
print('  - YOY %: Year-over-year (2025 vs 2024)')
print('  - LM %: Empty (not applicable for totals)')
print('=' * 80)
