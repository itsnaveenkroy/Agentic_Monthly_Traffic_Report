from openpyxl import load_workbook

# Check OUTPUT file
output_path = 'data/output_report.xlsx'
wb = load_workbook(output_path)
ws = wb.active

print('=' * 70)
print('TOTAL ROW CHECK')
print('=' * 70)
print(f'Row 23 YOY: {ws.cell(23, 6).value}')
print(f'Row 23 LM: {ws.cell(23, 7).value}')

print('\n' + '=' * 70)
print('SUMMARY FORMATTING CHECK')
print('=' * 70)

summary_cell = ws.cell(11, 9)
print(f'\nCell I11 value: {str(summary_cell.value)[:100]}...')
print(f'Font color: {summary_cell.font.color}')
print(f'Font color type: {type(summary_cell.font.color)}')
if summary_cell.font.color:
    print(f'Font color RGB: {summary_cell.font.color.rgb}')

# Check more sections
print('\n' + '=' * 70)
print('CHECKING MULTIPLE SECTION SUMMARIES FOR COLOR')
print('=' * 70)

section_rows = [11, 27, 44, 60, 76]  # First 5 sections
for row in section_rows:
    cell = ws.cell(row, 9)
    text = str(cell.value)[:60] if cell.value else 'None'
    color = cell.font.color.rgb if cell.font.color else 'None'
    has_declined = 'declined' in text.lower()
    has_upward = 'upward' in text.lower()
    print(f'\nRow {row}: declined={has_declined}, upward={has_upward}')
    print(f'  Text: {text}...')
    print(f'  Color: {color}')
