"""
Excel Writer Agent - LangGraph Node
Writes calculated metrics and summaries back to Excel.
"""

import pandas as pd
from typing import Dict, Any
import sys
import os

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.excel_utils import write_summary_to_excel, format_percentage_for_excel


class ExcelWriterAgent:
    """
    Agent responsible for writing results back to Excel workbook.
    """
    
    def __init__(self):
        """Initialize the Excel Writer Agent."""
        pass
    
    def find_yoy_lm_columns(self, worksheet, header_row: int) -> Dict[str, int]:
        """
        Find the column indices for YOY and LM columns.
        
        Args:
            worksheet: openpyxl worksheet object
            header_row: Row number containing column headers
            
        Returns:
            Dictionary with column indices
        """
        column_indices = {'yoy': None, 'lm': None, 'summary': None}
        
        # Read header row
        for col_idx, cell in enumerate(worksheet[header_row], start=1):
            if cell.value:
                cell_str = str(cell.value).lower()
                
                if 'yoy' in cell_str and '2024' in cell_str and '2025' in cell_str:
                    column_indices['yoy'] = col_idx
                elif 'lm' in cell_str and '2025' in cell_str:
                    column_indices['lm'] = col_idx
        
        # Summary column is typically 1-2 columns after LM
        if column_indices['lm']:
            column_indices['summary'] = column_indices['lm'] + 2
        elif column_indices['yoy']:
            column_indices['summary'] = column_indices['yoy'] + 3
        else:
            # Default to column H (8)
            column_indices['summary'] = 8
        
        return column_indices
    
    def write_metrics_to_excel(self, worksheet, df: pd.DataFrame, 
                               data_start_row: int, column_indices: Dict[str, int],
                               section_info: Dict[str, Any]):
        """
        FIX 5: Write YOY and LM metrics to Excel worksheet with debug logging.
        
        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame with calculated metrics
            data_start_row: Starting row for data
            column_indices: Dictionary with column indices
            section_info: Section metadata for finding Total/% Change rows
        """
        # Identify metric columns in DataFrame
        yoy_col_name = None
        lm_col_name = None
        month_col_name = df.columns[0] if len(df.columns) > 0 else None
        
        for col in df.columns:
            col_str = str(col).lower()
            if 'yoy' in col_str and ('2024' in col_str or '2025' in col_str):
                yoy_col_name = col
            elif 'lm' in col_str and '2025' in col_str:
                lm_col_name = col
        
        print(f"  [DEBUG] Writing metrics: YOY col='{yoy_col_name}', LM col='{lm_col_name}'")
        
        # Write values row by row with debug output
        rows_written = 0
        for df_idx, row in df.iterrows():
            excel_row = data_start_row + df_idx
            month_name = row[month_col_name] if month_col_name else f"Row {df_idx}"
            
            # Write YOY value
            if yoy_col_name and column_indices['yoy']:
                yoy_value = row[yoy_col_name]
                if pd.notna(yoy_value):
                    cell = worksheet.cell(row=excel_row, column=column_indices['yoy'])
                    cell.value = format_percentage_for_excel(yoy_value)
                    rows_written += 1
                    print(f"  [DEBUG] Writing YOY for row {excel_row} (Month={month_name}): {yoy_value:.2f}%")
            
            # Write LM value
            if lm_col_name and column_indices['lm']:
                lm_value = row[lm_col_name]
                if pd.notna(lm_value):
                    cell = worksheet.cell(row=excel_row, column=column_indices['lm'])
                    cell.value = format_percentage_for_excel(lm_value)
                    print(f"  [DEBUG] Writing LM for row {excel_row} (Month={month_name}): {lm_value:.2f}%")
        
        print(f"  [DEBUG] Total rows with metrics written: {rows_written}")
        
        # Calculate and write Total row with actual sum values
        self.write_total_row(worksheet, df, section_info)
        
        # Calculate and write % Change row
        input_path = section_info.get('input_path')
        self.write_percent_change_row(worksheet, df, section_info, column_indices, input_path)
    
    def write_total_row(self, worksheet, df: pd.DataFrame, section_info: Dict[str, Any]):
        """
        Calculate and write Total row with actual sum values.
        The Total row should show sum of all month values, not formulas.
        """
        # Find Total row
        data_end_row = section_info['data_end_row']
        total_row_idx = None
        
        for row_idx in range(section_info['data_start_row'], data_end_row + 1):
            cell_value = worksheet.cell(row_idx, 2).value
            if cell_value and isinstance(cell_value, str) and 'total' in cell_value.lower() and 'visits' not in cell_value.lower():
                total_row_idx = row_idx
                break
        
        if not total_row_idx:
            return
        
        # Find year columns in DataFrame and calculate sums
        header_row = section_info['header_row']
        
        for col_idx in range(3, 20):
            header_value = worksheet.cell(header_row, col_idx).value
            if header_value:
                header_str = str(header_value).lower()
                # Only sum year columns, not YOY/LM columns
                if ('year' in header_str or '2023' in header_str or '2024' in header_str or '2025' in header_str) and \
                   'yoy' not in header_str and 'lm' not in header_str:
                    # Find matching column in DataFrame
                    for df_col in df.columns:
                        if str(header_value).strip() == str(df_col).strip():
                            # Calculate sum excluding Total row itself
                            df_without_total = df[~df.iloc[:, 0].astype(str).str.lower().str.contains('total', na=False)]
                            total_sum = pd.to_numeric(df_without_total[df_col], errors='coerce').sum()
                            
                            if pd.notna(total_sum) and total_sum != 0:
                                worksheet.cell(row=total_row_idx, column=col_idx).value = round(total_sum, 0)
                                print(f"  [DEBUG] Wrote Total for {header_value}: {total_sum:.0f}")
                            break
    
    def write_percent_change_row(self, worksheet, df: pd.DataFrame,
                                  section_info: Dict[str, Any], column_indices: Dict[str, int],
                                  input_path: str = None):
        """
        IMPORTANT: % Change row behavior (REPORTING REQUIREMENT)
        
        % Change row uses first numeric column (Year-2023) as REFERENCE baseline.
        Year-2025 uses Year-2024 as reference (matching YOY logic).
        
        Column-wise behavior:
        - Month column: EMPTY
        - Year-2023 (first numeric): EMPTY (reference baseline - 0% change from itself)
        - Year-2024: (Total_2024 - Total_2023) / Total_2023 * 100
        - Year-2025: (Jan-Aug_2025 - Jan-Aug_2024) / Jan-Aug_2024 * 100 + " (till Aug)"
        - YOY %: (Jan-Aug_2025 - Jan-Aug_2024) / Jan-Aug_2024 * 100 + " (till Aug)"
        - LM %: EMPTY
        
        This % Change behavior is intentional. First column is a reference baseline.
        """
        # Find Total and % Change rows in the Excel sheet
        data_end_row = section_info['data_end_row']
        total_row_idx = None
        change_row_idx = None
        
        for row_idx in range(section_info['data_start_row'], data_end_row + 1):
            cell_value = worksheet.cell(row_idx, 2).value  # Check column B (Month column)
            if cell_value and isinstance(cell_value, str):
                cell_lower = cell_value.lower()
                if 'total' in cell_lower and 'visits' not in cell_lower:
                    total_row_idx = row_idx
                    print(f"  [DEBUG] Found Total row at {row_idx}")
                elif '% change' in cell_lower or '%change' in cell_lower:
                    change_row_idx = row_idx
                    print(f"  [DEBUG] Found % Change row at {row_idx}")
        
        if not total_row_idx or not change_row_idx:
            print(f"  [DEBUG] Missing rows - Total: {total_row_idx}, % Change: {change_row_idx}")
            return
        
        # STEP 1: Clear ALL columns in % Change row first (enforce empty state)
        for col_idx in range(2, 20):  # Columns B through S
            worksheet.cell(row=change_row_idx, column=col_idx).value = None
        
        # Add "% Change" label in column B (Month column)
        worksheet.cell(row=change_row_idx, column=2).value = "% Change"
        print(f"  [DEBUG] Cleared all % Change row cells and added label")
        
        # STEP 2: Find all year columns (2023, 2024, 2025)
        header_row = section_info['header_row']
        year_columns = []  # List of (col_idx, year)
        
        for col_idx in range(3, 20):
            header_value = worksheet.cell(header_row, col_idx).value
            if header_value:
                header_str = str(header_value).lower()
                if 'yoy' not in header_str and 'lm' not in header_str and '% change' not in header_str:
                    if '2023' in header_str:
                        year_columns.append((col_idx, '2023'))
                    elif '2024' in header_str:
                        year_columns.append((col_idx, '2024'))
                    elif '2025' in header_str:
                        year_columns.append((col_idx, '2025'))
        
        year_columns.sort(key=lambda x: x[1])  # Sort by year
        print(f"  [DEBUG] Found year columns: {[(y, c) for c, y in year_columns]}")
        
        if len(year_columns) < 2:
            print(f"  [DEBUG] Not enough year columns for % Change calculation")
            return
        
        # STEP 3: Load workbook with data_only=True to get values
        from openpyxl import load_workbook
        
        ws_data_only = None
        if input_path:
            try:
                wb_data_only = load_workbook(input_path, data_only=True)
                ws_data_only = wb_data_only[worksheet.title]
                print(f"  [DEBUG] Loaded data_only workbook for values")
            except Exception as e:
                print(f"  [DEBUG] Could not load data_only workbook: {e}")
        
        # STEP 4: Get Total values AND Jan-Aug values for each year
        total_values = {}
        jan_aug_values = {}
        
        # Define months Jan through Aug (rows to include)
        jan_aug_months = ['jan', 'feb', 'march', 'april', 'may', 'june', 'july', 'aug']
        
        for col_idx, year in year_columns:
            # Get Total value
            total_val = None
            if ws_data_only:
                total_val = ws_data_only.cell(total_row_idx, col_idx).value
                try:
                    total_val = float(total_val) if total_val is not None else None
                except:
                    total_val = None
            
            # Fallback: Sum from DataFrame
            if total_val is None:
                for col_name in df.columns:
                    if year in str(col_name).lower() and 'yoy' not in str(col_name).lower() and 'lm' not in str(col_name).lower():
                        total_val = pd.to_numeric(df[col_name], errors='coerce').sum()
                        break
            
            if total_val and total_val > 0:
                total_values[year] = total_val
                print(f"  [DEBUG] Total {year}: {total_val}")
            
            # Calculate Jan-Aug sum for this year
            jan_aug_sum = 0
            for row_idx in range(section_info['data_start_row'], total_row_idx):
                month_cell = worksheet.cell(row_idx, 2).value
                if month_cell and isinstance(month_cell, str):
                    month_lower = month_cell.strip().lower()
                    if any(m in month_lower for m in jan_aug_months):
                        # Get value from this row for the year column
                        if ws_data_only:
                            val = ws_data_only.cell(row_idx, col_idx).value
                        else:
                            val = worksheet.cell(row_idx, col_idx).value
                        
                        try:
                            val = float(val) if val is not None else 0
                            jan_aug_sum += val
                        except:
                            pass
            
            if jan_aug_sum > 0:
                jan_aug_values[year] = jan_aug_sum
                print(f"  [DEBUG] Jan-Aug {year}: {jan_aug_sum}")
        
        # STEP 5: Calculate % Change using FIRST year as reference
        if len(total_values) < 2:
            print(f"  [DEBUG] Not enough total values for % Change calculation")
            return
        
        reference_year = year_columns[0][1]  # First year is reference
        reference_total = total_values.get(reference_year)
        reference_jan_aug = jan_aug_values.get(reference_year)
        
        if not reference_total or reference_total <= 0:
            print(f"  [DEBUG] Invalid reference total for {reference_year}: {reference_total}")
            return
        
        print(f"  [DEBUG] Using {reference_year} as reference: Total={reference_total}, Jan-Aug={reference_jan_aug}")
        
        # Calculate % Change for each subsequent year vs reference
        for col_idx, year in year_columns:
            if year == reference_year:
                # Reference year stays empty (0% change from itself)
                continue
            
            # For Year-2025, use Jan-Aug comparison with Year-2024 (NOT reference year)
            if year == '2025' and '2024' in jan_aug_values:
                jan_aug_2024 = jan_aug_values.get('2024')
                current_jan_aug = jan_aug_values.get(year)
                if jan_aug_2024 and jan_aug_2024 > 0 and current_jan_aug is not None:
                    pct_change = ((current_jan_aug - jan_aug_2024) / jan_aug_2024) * 100
                    worksheet.cell(row=change_row_idx, column=col_idx).value = f"{pct_change:.2f}% (till Aug)"
                    print(f"  [DEBUG] % Change for {year} (Jan-Aug): {pct_change:.2f}% (vs 2024 Jan-Aug)")
                else:
                    print(f"  [DEBUG] No Jan-Aug value for {year} or 2024, skipping")
            else:
                # For other years (2024), use total comparison
                current_total = total_values.get(year)
                if current_total is not None:
                    pct_change = ((current_total - reference_total) / reference_total) * 100
                    worksheet.cell(row=change_row_idx, column=col_idx).value = format_percentage_for_excel(pct_change)
                    print(f"  [DEBUG] % Change for {year}: {pct_change:.2f}% (vs {reference_year})")
                else:
                    print(f"  [DEBUG] No total value for {year}, skipping")
        
        # STEP 6: YOY column in % Change row should remain EMPTY (calculation now in Year-2025 column)
        # This avoids duplication since Year-2025 now uses the same 2024 reference logic
        print(f"  [DEBUG] YOY column in % Change row left empty (calculation moved to Year-2025 column)")
    
    def write_summary_to_section(self, worksheet, summary_text: str,
                                 data_start_row: int, data_end_row: int,
                                 summary_column_idx: int):
        """
        Write executive summary with proper formatting, merging, width, and keyword coloring.
        - Entire summary colored #22577A if contains 'declined'
        - Entire summary colored GREEN if contains 'upward'
        
        This formatting is presentation-only. Do not modify logic.
        
        Args:
            worksheet: openpyxl worksheet object
            summary_text: Summary text to write
            data_start_row: Starting row for summary
            data_end_row: Ending row for summary
            summary_column_idx: Column index for summary
        """
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter
        import re
        
        summary_column_letter = get_column_letter(summary_column_idx)
        
        # Set column width for readability
        worksheet.column_dimensions[summary_column_letter].width = 60
        
        # Add header "Summary / Insights" one row above the summary block
        header_row = data_start_row - 1
        header_cell = worksheet.cell(row=header_row, column=summary_column_idx)
        header_cell.value = "Summary / Insights :"
        header_cell.font = Font(name='Century Gothic', size=12, bold=True)
        header_cell.alignment = Alignment(horizontal='left', vertical='top')
        print(f"  [DEBUG] Added header 'Summary / Insights' at {summary_column_letter}{header_row}")
        
        # Determine font color based on keywords (before merging)
        has_declined = bool(re.search(r'\bdeclined?\b', summary_text, re.IGNORECASE))
        has_upward = bool(re.search(r'\bupward\b', summary_text, re.IGNORECASE))
        
        print(f"  [DEBUG] Summary text preview: {summary_text[:100]}...")
        print(f"  [DEBUG] Keyword detection - declined: {has_declined}, upward: {has_upward}")
        
        # Determine the font color to use
        # Priority: upward trend takes precedence when both keywords exist (indicates overall positive despite mentioning declines)
        if has_upward:
            font_color = '00B050'  # Green text for upward trend
            color_name = "GREEN"
        elif has_declined:
            font_color = '22577A'  # Blue color for decline
            color_name = "BLUE (#22577A)"
        else:
            font_color = '000000'  # Default black
            color_name = "BLACK"
        
        print(f"  [DEBUG] Selected color: {color_name} (hex: {font_color})")
        
        # This formatting is presentation-only. Do not modify logic.
        # Define light gray border
        light_gray_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Merge cells vertically for the summary block
        merge_range = f"{summary_column_letter}{data_start_row}:{summary_column_letter}{data_end_row}"
        try:
            worksheet.merge_cells(merge_range)
            print(f"  [DEBUG] Merged summary cells: {merge_range}")
        except Exception as e:
            print(f"  [DEBUG] Could not merge cells {merge_range}: {e}")
        
        # Write summary text and apply formatting to the first cell of merged range
        summary_cell = worksheet.cell(row=data_start_row, column=summary_column_idx)
        summary_cell.value = summary_text
        summary_cell.border = light_gray_border
        summary_cell.font = Font(name='Century Gothic', size=12, color=font_color)
        summary_cell.alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrap_text=True  # Enable text wrapping
        )
        
        print(f"  [DEBUG] Applied {color_name} formatting and light gray border to summary block {merge_range}")
        print(f"  [DEBUG] Wrote summary to {summary_column_letter}{data_start_row} with text wrapping")
    
    def execute(self, state: Dict[str, Any]) -> Dict[str, Any]:
        """
        Write results back to Excel workbook.
        
        Args:
            state: Current LangGraph state
            
        Returns:
            Updated state
        """
        print(f"\n{'='*60}")
        print("EXCEL WRITER AGENT")
        print(f"{'='*60}")
        
        worksheet = state['worksheet']
        section_name = state['section_name']
        df = state['calculated_metrics']
        summary_text = state['summary_text']
        section_info = state['section_info']
        
        print(f"[+] Writing results for: {section_name}")
        
        # Get row information
        header_row = section_info['header_row']  # Headers are in same row as section name
        data_start_row = section_info['data_start_row']
        data_end_row = data_start_row + len(df) - 1
        
        # Find column indices
        column_indices = self.find_yoy_lm_columns(worksheet, header_row)
        
        print(f"[+] Writing metrics to rows {data_start_row}-{data_end_row}")
        
        # Write metrics (FIX 5: pass section_info for % Change row)
        self.write_metrics_to_excel(worksheet, df, data_start_row, column_indices, section_info)
        
        print(f"[+] Metrics written to YOY column {column_indices['yoy']}, LM column {column_indices['lm']}")
        
        # Write summary
        if column_indices['summary']:
            print(f"[+] Writing summary to column {column_indices['summary']}")
            self.write_summary_to_section(
                worksheet,
                summary_text,
                data_start_row,
                data_end_row,
                column_indices['summary']
            )
            print(f"[+] Summary written successfully")
        
        print(f"[+] Excel writing complete for {section_name}")
        print(f"{'='*60}\n")
        
        return state


def create_excel_writer_node():
    """
    Factory function to create Excel Writer node for LangGraph.
    
    Returns:
        Callable node function
    """
    agent = ExcelWriterAgent()
    return agent.execute
