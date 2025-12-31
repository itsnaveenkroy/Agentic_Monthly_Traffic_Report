"""
Excel Reader Agent - LangGraph Node
Loads the Excel workbook and worksheet into the shared state.
"""

from openpyxl import load_workbook
from typing import Dict, Any
import os


class ExcelReaderAgent:
    """
    Agent responsible for loading Excel file and preparing it for processing.
    """
    
    def __init__(self, input_path: str):
        """
        Initialize the Excel Reader Agent.
        
        Args:
            input_path: Path to input Excel file
        """
        self.input_path = input_path
    
    def execute(self, state: Dict[str, Any]) -> Dict[str, Any]:
        """
        Load Excel workbook and worksheet.
        
        Args:
            state: Current LangGraph state
            
        Returns:
            Updated state with workbook and worksheet
        """
        print(f"\n{'='*60}")
        print("EXCEL READER AGENT")
        print(f"{'='*60}")
        
        # Validate file exists
        if not os.path.exists(self.input_path):
            raise FileNotFoundError(f"Input Excel file not found: {self.input_path}")
        
        print(f"[+] Loading workbook: {self.input_path}")
        
        # Load workbook with openpyxl
        workbook = load_workbook(self.input_path)
        
        # Get first worksheet
        worksheet = workbook.active
        sheet_name = worksheet.title
        
        print(f"[+] Loaded worksheet: '{sheet_name}'")
        print(f"[+] Worksheet dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns")
        
        # Update state
        state['workbook'] = workbook
        state['worksheet'] = worksheet
        state['sheet_name'] = sheet_name
        state['input_path'] = self.input_path  # Store path for data_only reload
        
        print(f"[+] Excel workbook loaded successfully")
        print(f"{'='*60}\n")
        
        return state


def create_excel_reader_node(input_path: str):
    """
    Factory function to create Excel Reader node for LangGraph.
    
    Args:
        input_path: Path to input Excel file
        
    Returns:
        Callable node function
    """
    agent = ExcelReaderAgent(input_path)
    return agent.execute
