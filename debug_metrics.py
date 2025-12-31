import pandas as pd
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('Traffic_status.xlsx')
ws = wb.active

# Extract first section data (Total Visits)
headers = []
for cell in list(ws[10])[1:]:  # Row 10, skip column A
    if cell.value:
        headers.append(str(cell.value).strip())
    else:
        break

print(f"Headers: {headers[:6]}")

# Extract data rows 11-23
data_rows = []
for row_idx in range(11, 24):  # Rows 11-23
    row_data = [cell.value for cell in list(ws[row_idx])[1:len(headers)+1]]
    data_rows.append(row_data)
    print(f"Row {row_idx}: {row_data[:6]}")

# Create DataFrame
df = pd.DataFrame(data_rows, columns=headers[:len(data_rows[0])])
print(f"\nDataFrame shape: {df.shape}")
print(f"Month column values: {df.iloc[:, 0].tolist()}")

# Check Total row
total_row = df[df.iloc[:, 0].astype(str).str.lower().str.contains('total', na=False)]
print(f"\nTotal row found: {len(total_row)} rows")
if len(total_row) > 0:
    print(f"Total row index: {total_row.index.tolist()}")
    print(f"Total row data: {total_row.iloc[0, :6].tolist()}")
